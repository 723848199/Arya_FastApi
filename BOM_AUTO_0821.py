# !/bin/env python
# -*- coding:utf-8 -*-
# 新增需求：每次执行更新New sheet，不新增
from bom_e_auto import add
'''*
今日工作
根据的新需求实现BOM物料表自动化操作单元格：
需求：操作表sheet2，自动完成单元格修改需求，sheet1为原表，sheet2为操作需求表，（需求表逐行从第二行按顺序执行，第一行是标题），new sheet是新表

逻辑细节：级别满足1234，sheet2物料表编码和sheet1（原表）一致，判断操作是否为ADD、DEL，分别执行ADD或DEL操作，.ADD：1转字符串2.split切割3.循环判断，not append,4join拼接；5.len读位号个数，.DEL：1转字符串2.split切割3.循环判断，4join拼接；5.len读位号个数，not move,4join拼接

实现版本：打包成EXE，和excel在同级目录下，将EXE拖到同级目录下的终端 ，回车执行
'''

import openpyxl as op

# import xlrd as xr
import xlwt as xw

path = './bomexcel_auto.xlsx'
# 2.打开Excel文件
wb = op.load_workbook(path)

# 3.获取指定Sheet对象
sheet1 = wb['Sheet1']

# 4.复制 Sheet 对象
sheet3 = wb.copy_worksheet(sheet1)

# 5. 重新命名 sheet 对象
sheet3.title = 'new_excel'

or_excel = wb["Sheet1"]
operate_excel = wb["Sheet2"]
new_excel = wb["new_excel"]

print("Sheet2操作表单 有" + f'{operate_excel.max_row}' + "行")
print("Sheet2操作表单 有" + f'{operate_excel.max_column}' + "列")

# 开始逻辑判断
for count in range(operate_excel.max_row - 1):
    if operate_excel[f"A{count + 2}"].value in [1, 2, 3, 4]:
        if new_excel[f"E{count + 2}"].value == operate_excel[f"B{count + 2}"].value:
            print(f'{count + 1}'+'、执行' + f'{new_excel[f"E{count + 2}"].value}' + '物料编码')

            if operate_excel[f"C{count + 2}"].value == 'del':
                print(" 现在对操作表第" + f'{count + 2}' + "行"+str(
                    new_excel[f"E{count + 2}"].value) + "物料编码进行 删除操作，操作内容是：\'" + str(
                    operate_excel[f"D{count + 2}"].value) + "\'========  开始执行删除操作=========")
                #  级别==3or4,opr ==delete, 删除指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                new_excel_weihao = str(new_excel[f"P{count + 2}"].value)
                # print(new_excel_weihao)
                new_excel_weihao = new_excel_weihao.split(",")
                weihao_delete = operate_excel[f"E{count + 2}"].value.split(",")

                for value in weihao_delete:
                    if value in new_excel_weihao:
                        new_excel_weihao.remove(value)
                new_excel[f"P{count + 2}"] = ",".join(map(str, new_excel_weihao))
                # 菲菱子件用量 L
                new_excel[f"L{count + 2}"] = len(new_excel_weihao)
                '''
                # 6. 保存工作薄
                wb.save('./data.xlsx')
                '''
                print(" 对" + str(new_excel[f"E{count + 2}"].value) + "\'========  删除操作已完成=========")

                # print(or_excel[f"P{count + 2}"].value)
            elif operate_excel[f"C{count + 2}"].value == 'add':
                print(" 现在对操作表第" + f'{count + 2}' + "行"+ str(new_excel[f"E{count + 2}"].value) + "物料编码进行 增加操作，操作内容是：\'" + str(
                    operate_excel[f"D{count + 2}"].value) + "\'")
                #  级别==3or4,opr ==add, 增加指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                new_excel_weihao = str(new_excel[f"P{count + 2}"].value)
                # print(new_excel_weihao)

                new_excel_weihao = new_excel_weihao.split(",")
                weihao_add = operate_excel[f"E{count + 2}"].value.split(",")
                for value in weihao_add:
                    if value not in new_excel_weihao:
                        new_excel_weihao.append(value)
                        # print(weihao_or)
                #  用 join() 拼接 add 后的字段
                # or_excel[f"E{count + 2}"] = str.join(weihao_or)
                #  用 map() 拼接 add 后的字段
                new_excel[f"P{count + 2}"] = ",".join(map(str, new_excel_weihao))
                # 菲菱子件用量 L
                new_excel[f"L{count + 2}"] = len(new_excel_weihao)
                # 6. 保存工作薄
                wb.save('./data.xlsx')

                print(" 对" + str(new_excel[f"E{count + 2}"].value) + "\'======== 增加操作已完成=========")
        else:
            # 找到对应物料编码，替换整个单元格的数据值sub
            pass
    # elif operate_excel[f"A{count + 2}"].value in [3, 4]:
    #     #  级别==1or2, 替换单元格数据
    #     or_excel[f"P{count + 2}"] = operate_excel[f"E{count + 2}"].value
    #     print(or_excel[f"P{count + 2}"].value)
    count = +1

    # 6. 保存工作薄
    wb.save('./bomexcel_auto.xlsx')



