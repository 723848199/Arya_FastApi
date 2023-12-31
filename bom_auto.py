# !/bin/env python
# -*- coding:utf-8 -*-
import openpyxl as op

# import xlrd as xr
import xlwt as xw

path = './bomexcel_auto.xlsx'
# 2.打开Excel文件
wb = op.load_workbook(path)

# 3.获取指定Sheet对象
sheet1 = wb['原表']

# 获取工作表名字列表
sheet_names = wb.sheetnames
print(sheet_names)

# 判断 "Sheet1" 是否在工作表名字列表中，若tip_sheet不存在，新增 tip_sheet
if "新BOM表" in sheet_names:
    print(" 新BOM表 sheet1 已存在")
else:
    # 4.复制 Sheet 对象
    sheet3 = wb.copy_worksheet(sheet1)
    # 5. 重新命名 sheet3 对象
    sheet3.title = '新BOM表'
    # 保存
    wb.save('./bomexcel_auto.xlsx')

# 判断 "tip_sheet" 是否在工作表名字列表中，若tip_sheet不存在，新增 tip_sheet
if "tip_sheet" in sheet_names:
    print(" 温馨提示tip_sheet 已存在")
else:
    wb.create_sheet(title='温馨提示', index=0)
    wb.save('./bomexcel_auto.xlsx')


# 操作对象
or_excel = wb["原表"]
operate_excel = wb["操作指示表"]
new_excel = wb['新BOM表']
tip_sheet = wb["温馨提示"]

# tip行列提示
row_tip = "Sheet2操作表单 有" + f'{operate_excel.max_row}' + "行"
col_tip = "Sheet2操作表单 有" + f'{operate_excel.max_column}' + "列"
# print("Sheet2操作表单 有" + f'{operate_excel.max_row}' + "行")
# print("Sheet2操作表单 有" + f'{operate_excel.max_column}' + "列")
tip_sheet[f"A1"] = "此表为操作提示表：操作详情可对照此表核验、检查是否有遗漏操作"
tip_sheet[f"A2"] = row_tip
tip_sheet[f"A3"] = col_tip

# 开始逻辑判断
for count in range(operate_excel.max_row - 1):
    if operate_excel[f"A{count + 2}"].value in [1, 2, 3, 4]:
        if new_excel[f"E{count + 2}"].value == operate_excel[f"B{count + 2}"].value:
            print(f'{count + 1}' + '、执行' + f'{new_excel[f"E{count + 2}"].value}' + '物料编码')
            if operate_excel[f"C{count + 2}"].value == 'del':
                b = f'{count + 1}' + '、 对操作表第' + f'{count + 2}' + "行" + str(
                    new_excel[f"E{count + 2}"].value) + "物料编码进行 删除操作，操作内容是：\'" + str(
                    operate_excel[f"D{count + 2}"].value) + "\'========  开始执行删除操作==========  删除操作已完成========="
                # print(a)
                tip_sheet[f"A{count + 5}"] = b
                # print(" 现在对操作表第" + f'{count + 2}' + "行" + str(
                #     new_excel[f"E{count + 2}"].value) + "物料编码进行 删除操作，操作内容是：\'" + str(
                #     operate_excel[f"D{count + 2}"].value) + "\'========  开始执行删除操作=========")
                #  级别==3 or 4,opr ==delete, 删除指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                new_excel_weihao = str(new_excel[f"P{count + 2}"].value)
                # print(new_excel_weihao)
                new_excel_weihao = new_excel_weihao.split(",")
                weihao_delete = operate_excel[f"E{count + 2}"].value.split(",")

                for value in weihao_delete:
                    if value in new_excel_weihao:
                        new_excel_weihao.remove(value)
                new_excel[f"P{count + 2}"] = ",".join(map(str, new_excel_weihao))
                # 菲菱子件用量 L
                new_excel[f"L{count + 2}"] = len(new_excel_weihao)
                '''
                # 6. 保存工作薄
                wb.save('./data.xlsx')
                '''
                # print(" 对" + str(new_excel[f"E{count + 2}"].value) + "\'========  删除操作已完成=========")

                # print(or_excel[f"P{count + 2}"].value)
            elif operate_excel[f"C{count + 2}"].value == 'add':
                # print(" 现在对操作表第" + f'{count + 2}' + "行" + str(
                #     new_excel[f"E{count + 2}"].value) + "物料编码进行 增加操作，操作内容是：\'" + str(
                # operate_excel[f"D{count + 2}"].value) + "\'========  开始执行增加操作=========")
                a = f'{count + 1}' + '、 对操作表第' + f'{count + 2}' + "行" + str(
                    new_excel[f"E{count + 2}"].value) + "物料编码进行 增加操作，操作内容是：\'" + str(
                    operate_excel[f"D{count + 2}"].value) + "\'========  开始执行增加操作=========== 增加操作已完成========="
                # print(a)
                tip_sheet[f"A{count + 5}"] = a

                #  级别==3or4,opr ==add, 增加指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                new_excel_weihao = str(new_excel[f"P{count + 2}"].value)
                # print(new_excel_weihao)

                new_excel_weihao = new_excel_weihao.split(",")
                weihao_add = operate_excel[f"E{count + 2}"].value.split(",")
                for value in weihao_add:
                    if value not in new_excel_weihao:
                        new_excel_weihao.append(value)
                        # print(weihao_or)
                #  用 join() 拼接 add 后的字段
                # or_excel[f"E{count + 2}"] = str.join(weihao_or)
                #  用 map() 拼接 add 后的字段
                new_excel[f"P{count + 2}"] = ",".join(map(str, new_excel_weihao))
                # 菲菱子件用量 L
                new_excel[f"L{count + 2}"] = len(new_excel_weihao)
                # 6. 保存工作薄
                # wb.save('./data.xlsx')

                # print(" 对" + str(new_excel[f"E{count + 2}"].value) + "\'======== 增加操作已完成=========")
        else:
            # 找到对应物料编码，替换整个单元格的数据值sub
            pass
    # elif operate_excel[f"A{count + 2}"].value in [3, 4]:
    #     #  级别==1or2, 替换单元格数据
    #     or_excel[f"P{count + 2}"] = operate_excel[f"E{count + 2}"].value
    #     print(or_excel[f"P{count + 2}"].value)
    # 6. 保存工作薄
    wb.save('./bomexcel_auto.xlsx')


def add():
    print(" 对" + str(new_excel[f"E{count + 2}"].value) + "物料编码进行 增加操作，操作内容是：\'" + str(
        operate_excel[f"D{count + 2}"].value) + "\'")
    #  级别==3or4,opr ==add, 增加指定单元格数据
    #  分别对 位号org  和 opr 用逗号分割，转化成列表
    new_excel_weihao = str(new_excel[f"P{count + 2}"].value)
    # print(new_excel_weihao)

    new_excel_weihao = new_excel_weihao.split(",")
    weihao_add = operate_excel[f"E{count + 2}"].value.split(",")
    for value in weihao_add:
        if value not in new_excel_weihao:
            new_excel_weihao.append(value)
            # print(weihao_or)
    #  用 join() 拼接 add 后的字段
    # or_excel[f"E{count + 2}"] = str.join(weihao_or)
    #  用 map() 拼接 add 后的字段
    new_excel[f"P{count + 2}"] = ",".join(map(str, new_excel_weihao))
    # 菲菱子件用量 L
    new_excel[f"L{count + 2}"] = len(new_excel_weihao)
    # 6. 保存工作薄
    wb.save('./data.xlsx')

    print(" 对" + str(new_excel[f"E{count + 2}"].value) + "物料编码进行增加操作，操作内容是：\'" + str(
        operate_excel[f"D{count + 2}"].value) + "\'======== 增加操作已完成=========")
