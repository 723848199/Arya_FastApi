import openpyxl as op
# import xlrd as xr
import xlwt as xw

# 打开Excel文件
wb = op.load_workbook("data.xlsx")

# 获取指定Sheet
or_excel = wb["Sheet1"]
operate_excel = wb["Sheet2"]
print(operate_excel.max_row)
# print(or_excel.max_column)
for count in range(operate_excel.max_row - 1):
    # print(or_excel["E2"].value)
    # print(operate_excel["B2"].value)
    # 添加新的Sheet
    new_sheet = wb.create_sheet("Sheet2")

    if operate_excel[f"A{count + 2}"].value in [3, 4]:
        if or_excel[f"E{count + 2}"].value == operate_excel[f"B{count + 2}"].value:
            print(or_excel[f"E{count + 2}"].value)
            if operate_excel[f"C{count + 2}"].value == 'delete':
                print(" 对" + str(or_excel[f"E{count + 2}"].value) + "物料编码进行 删除操作，操作内容是：\'" + str(
                        operate_excel[f"D{count + 2}"].value) + "\'")
                #  级别==3or4,opr ==delete, 删除指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                weihao = (or_excel[f"P{count + 2}"].value).split(",")
                weihao_delete = operate_excel[f"E{count + 2}"].value.split(",")
                for value in weihao_delete:
                    if value in weihao:
                        weihao.remove(value)
                print(" 对" + str(or_excel[f"E{count + 2}"].value) + "物料编码进行删除操作，操作内容是：\'" + str(
                        operate_excel[f"D{count + 2}"].value) + "\'========  删除操作已完成=========")

                # print(or_excel[f"P{count + 2}"].value)
            elif operate_excel[f"C{count + 2}"].value == 'add':
                print(" 对" + str(or_excel[f"E{count + 2}"].value) + "物料编码进行 增加操作，操作内容是：\'" + str(
                        operate_excel[f"D{count + 2}"].value) + "\'")
                #  级别==3or4,opr ==add, 增加指定单元格数据
                #  分别对 位号org  和 opr 用逗号分割，转化成列表
                weihao_or = (or_excel[f"P{count + 2}"].value).split(",")
                weihao_add = operate_excel[f"E{count + 2}"].value.split(",")
                for value in weihao_add:
                    if value not in weihao_or:
                        weihao_or.append(value)
                        # print(weihao_or)
                #  用 join() 拼接 add 后的字段
                # or_excel[f"E{count + 2}"] = str.join(weihao_or)
                #  用 map() 拼接 add 后的字段
                or_excel[f"E{count + 2}"] = ",".join(map(str, weihao_or))
                print(or_excel[f"E{count + 2}"].value)
                print(" 对" + str(or_excel[f"E{count + 2}"].value) + "物料编码进行增加操作，操作内容是：\'" + str(
                        operate_excel[f"D{count + 2}"].value) + "\'======== 增加操作已完成=========")
        else:
            # 找到对应物料编码，替换整个单元格的数据值sub
            pass
    elif operate_excel[f"A{count + 2}"].value in [3, 4]:
        #  级别==1or2, 替换单元格数据
        or_excel[f"P{count + 2}"] = operate_excel[f"E{count + 2}"].value
        print(or_excel[f"P{count + 2}"].value)
        # 保存修改后的Excel文件
    # wb.save("data_modified.xlsx")
